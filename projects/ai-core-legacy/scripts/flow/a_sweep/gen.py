#!/usr/bin/env python3

# -----------------------------------------------------------------------------
# Author: Simone Machetti
#
# Read doc/data/a_sweep/results.xlsx and write the A-magnitude sweep charts
# under doc/charts/a_sweep/. See scripts/flow/a_sweep/README.md for formulas.
# -----------------------------------------------------------------------------

from __future__ import annotations

import os
import sys
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

A_VALUES = (0, 1, 2, 3, 4, 5, 6, 7)
CONFIGS = ("config_1", "config_2", "config_3", "config_4")


def repo_root() -> Path:
    repo_home = os.environ.get("REPO_HOME")
    if not repo_home:
        sys.exit("REPO_HOME is not set; run `source sourceme.sh` first.")
    return Path(repo_home) / "projects" / "ai-core-legacy"


def load_improvement_pct(wb, sheet: str) -> np.ndarray:
    """Return improvement-% column (last column) indexed by A."""
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # header: A | baseline_mw | square_mw | delta_mw | improvement_pct
    data = np.zeros(len(A_VALUES))
    for r in rows[1:]:
        a = int(r[0])
        data[A_VALUES.index(a)] = float(r[4])
    return data


def heatmap_row(data: np.ndarray, title: str, out: Path) -> None:
    """1-D heatmap: a single colored row indexed by A."""
    vmax = float(np.max(np.abs(data)))
    fig, ax = plt.subplots(figsize=(9, 2.5))
    im = ax.imshow(data.reshape(1, -1), cmap="RdYlGn",
                   vmin=-vmax, vmax=vmax, aspect="auto")
    ax.set_xticks(np.arange(len(A_VALUES)))
    ax.set_xticklabels(A_VALUES)
    ax.set_yticks([0]); ax.set_yticklabels(["MAX_VAL_B = 127"])
    ax.set_xlabel("MAX_VAL_A")
    ax.set_title(title)
    for j, v in enumerate(data):
        color = "white" if abs(v) > vmax * 0.55 else "black"
        ax.text(j, 0, f"{v:+.2f}", ha="center", va="center",
                fontsize=9, color=color)
    cbar = fig.colorbar(im, ax=ax, orientation="vertical", fraction=0.04, pad=0.04)
    cbar.set_label("Improvement (%)  —  positive = Square wins")
    plt.tight_layout()
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"  wrote {out}")


def plot_lines(all_imp: dict[str, np.ndarray], out: Path) -> None:
    """All four configurations on a single line plot."""
    colors  = {"config_1": "#d73027", "config_2": "#006837",
               "config_3": "#1f77b4", "config_4": "#7d3c98"}
    markers = {"config_1": "o", "config_2": "s", "config_3": "^", "config_4": "D"}
    fig, ax = plt.subplots(figsize=(9, 5))
    for cfg in CONFIGS:
        ax.plot(A_VALUES, all_imp[cfg], color=colors[cfg], marker=markers[cfg],
                linewidth=2, markersize=8, label=cfg)
    ax.axhline(0, color="black", linewidth=0.8, linestyle=":", alpha=0.5)
    ax.set_xticks(A_VALUES)
    ax.set_xlabel("MAX_VAL_A")
    ax.set_ylabel("Improvement (%)  —  positive = Square wins")
    ax.set_title("A-magnitude sweep — Square vs Baseline at B = 127 (normal dist)")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="best")
    plt.tight_layout()
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"  wrote {out}")


CONFIG_TITLES = {
    "config_1": "Configuration 1 — Square vs Baseline (%) — PE arms only",
    "config_2": "Configuration 2 — Square vs Baseline (%) — with both alpha arms",
    "config_3": "Configuration 3 — Square vs Baseline (%) — 8x8 PE array",
    "config_4": "Configuration 4 — Square vs Baseline (%) — 8x16 PE array",
}


def main() -> int:
    root = repo_root()
    xlsx = root / "doc" / "data" / "a_sweep" / "results.xlsx"
    charts = root / "doc" / "charts" / "a_sweep"
    charts.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx, data_only=True)
    imp = {cfg: load_improvement_pct(wb, cfg) for cfg in CONFIGS}

    for cfg in CONFIGS:
        heatmap_row(imp[cfg], CONFIG_TITLES[cfg], charts / f"{cfg}.png")
    plot_lines(imp, charts / "improvement.png")
    return 0


if __name__ == "__main__":
    sys.exit(main())
