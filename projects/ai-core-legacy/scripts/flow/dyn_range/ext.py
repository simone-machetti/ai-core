#!/usr/bin/env python3

# -----------------------------------------------------------------------------
# Author: Simone Machetti
#
# Collate the 72 DPA reports from scripts/flow/dyn_range/run.py and
# write doc/data/dyn_range/results.xlsx (8 sheets).
# See scripts/flow/dyn_range/README.md.
# -----------------------------------------------------------------------------

from __future__ import annotations

import os
import sys
from pathlib import Path

from openpyxl import Workbook

A_VALUES = (1, 3, 5, 7)
B_VALUES = (1, 19, 37, 55, 73, 91, 109, 127)
SUFFIX = "dyn"


def repo_root() -> Path:
    repo_home = os.environ.get("REPO_HOME")
    if not repo_home:
        sys.exit("REPO_HOME is not set; run `source sourceme.sh` first.")
    return Path(repo_home) / "projects" / "ai-core-legacy"


def read_power_w(rpt: Path) -> float:
    for line in rpt.read_text().splitlines():
        if line.startswith("Total"):
            return float(line.split()[4])
    raise RuntimeError(f"No 'Total' row in {rpt}")


def to_uw(w: float) -> float:
    return w * 1e6


def load_2d(imp: Path, slug: str) -> dict[tuple[int, int], float]:
    out = {}
    for a in A_VALUES:
        for b in B_VALUES:
            rpt = imp / f"{slug}_{SUFFIX}_a{a}_b{b}_dpa" / "report" / "power_summary.rpt"
            out[(a, b)] = round(to_uw(read_power_w(rpt)), 1)
    return out


def load_1d(imp: Path, slug: str) -> dict[int, float]:
    out = {}
    for a in A_VALUES:
        rpt = imp / f"{slug}_{SUFFIX}_a{a}_dpa" / "report" / "power_summary.rpt"
        out[a] = round(to_uw(read_power_w(rpt)), 1)
    return out


def write_power_2d(wb: Workbook, name: str, data: dict[tuple[int, int], float]) -> None:
    ws = wb.create_sheet(name)
    ws.append(["A \\ B"] + list(B_VALUES))
    for a in A_VALUES:
        ws.append([a] + [data[(a, b)] for b in B_VALUES])


def write_power_1d(wb: Workbook, name: str, data: dict[int, float]) -> None:
    ws = wb.create_sheet(name)
    ws.append(["A", "power_uw"])
    for a in A_VALUES:
        ws.append([a, data[a]])


def write_config(wb: Workbook, name: str,
                      bas, sqr_sc, a_sqr, a_sum,
                      bas_mul: int, sqr_mul: int,
                      alpha_sqr_mul: int, alpha_sum_mul: int) -> None:
    ws = wb.create_sheet(name)
    ws.append(["A \\ B"] + list(B_VALUES))
    for a in A_VALUES:
        row = [a]
        for b in B_VALUES:
            base_total = bas[(a, b)] * bas_mul
            sqr_total = (sqr_sc[(a, b)] * sqr_mul
                         + a_sqr[a] * alpha_sqr_mul
                         + a_sum[a] * alpha_sum_mul)
            imp = 100.0 * (base_total - sqr_total) / base_total
            row.append(round(imp, 2))
        ws.append(row)


def main() -> int:
    root = repo_root()
    imp = root / "imp"
    out = root / "doc" / "data" / "dyn_range" / "results.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    bas    = load_2d(imp, "bas_4x8")
    sqr_sc = load_2d(imp, "sqr_4x8_sc")
    a_sum  = load_1d(imp, "sqr_4x8_alpha_sum")
    a_sqr  = load_1d(imp, "sqr_4x8_alpha_sqr")

    wb = Workbook()
    wb.remove(wb.active)
    write_power_2d(wb, "power_bas_4x8",    bas)
    write_power_2d(wb, "power_sqr_4x8_sc", sqr_sc)
    write_power_1d(wb, "power_alpha_sum",  a_sum)
    write_power_1d(wb, "power_alpha_sqr",  a_sqr)
    write_config(wb, "config_1", bas, sqr_sc, a_sqr, a_sum,
                 bas_mul=256, sqr_mul=256, alpha_sqr_mul=64, alpha_sum_mul=0)
    write_config(wb, "config_2", bas, sqr_sc, a_sqr, a_sum,
                 bas_mul=256, sqr_mul=256, alpha_sqr_mul=64, alpha_sum_mul=48)
    write_config(wb, "config_3", bas, sqr_sc, a_sqr, a_sum,
                 bas_mul=64,  sqr_mul=64,  alpha_sqr_mul=32, alpha_sum_mul=24)
    write_config(wb, "config_4", bas, sqr_sc, a_sqr, a_sum,
                 bas_mul=128, sqr_mul=128, alpha_sqr_mul=48, alpha_sum_mul=40)

    wb.save(out)
    print(f"  wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
