#!/usr/bin/env python3

# -----------------------------------------------------------------------------
# Author: Simone Machetti
#
# Read doc/data/regres/results.xlsx and write the eight regression charts under
# doc/charts/regres/. See scripts/flow/regres/README.md for definitions of the
# AI-core composition rules.
# -----------------------------------------------------------------------------

from __future__ import annotations

import os
import sys
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from openpyxl import load_workbook

C_BAR     = "#005f73"
C_BAS_4   = "#97665b"
C_SQR_4   = "#005f73"
C_ALPHA_4 = "#97bdc5"


def repo_root() -> Path:
    code_home = os.environ.get("CODE_HOME")
    if not code_home:
        sys.exit("CODE_HOME is not set; run `source sourceme.sh` first.")
    return Path(code_home) / "ai-core" / "projects" / "ai-core"


def load_xlsx(path: Path) -> dict[str, dict[str, float]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["data"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    out: dict[str, dict[str, float]] = {}
    for design, area_um2, freq_mhz, power_mw in rows:
        if design is None:
            continue
        out[design] = {
            "area_um2": float(area_um2),
            "freq_ghz": float(freq_mhz) / 1000.0,
            "power_mw": float(power_mw),
        }
    return out


def plot_pe_level(d: dict, key: str, title: str, ylabel: str, fmt: str, out: Path) -> None:
    names = list(d.keys())
    values = np.array([d[n][key] for n in names])
    x = np.arange(len(names))

    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(x, values, 0.5, color=C_BAR)
    for i, v in enumerate(values):
        ax.text(x[i], v * 1.01, fmt.format(v=v), ha="center", va="bottom", fontsize=9)
    ax.set_ylim(0, values.max() * 1.20)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=15, ha="right")
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    plt.tight_layout()
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"  wrote {out}")


def plot_ai_core_stacked(d: dict, key: str, n: int, title: str, ylabel: str,
                         baseline_fmt: str, include_alpha_sum: bool, out: Path) -> None:
    bas_4x8       = n*n * d["Baseline 4x8"][key]
    bas_8x8       = n*n * d["Baseline 8x8"][key]
    sqr_4x8_pe    = n*n * d["Square 4x8 SC"][key]
    if include_alpha_sum:
        sqr_4x8_alpha = n * (4 * d["Square 4x8 Alpha Squared"][key]
                             + 3 * d["Square 4x8 Alpha"][key])
    else:
        sqr_4x8_alpha = n * (4 * d["Square 4x8 Alpha Squared"][key])
    sqr_4x8       = sqr_4x8_pe + sqr_4x8_alpha
    sqr_8x8_pe    = n*n * d["Square 8x8"][key]
    sqr_8x8_alpha = n * (4 * d["Square 8x8 Alpha Squared"][key])
    sqr_8x8       = sqr_8x8_pe + sqr_8x8_alpha

    x = np.arange(4)
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(x[0], bas_4x8,       0.5, color=C_BAS_4)
    ax.bar(x[1], bas_8x8,       0.5, color=C_SQR_4)
    ax.bar(x[2], sqr_4x8_pe,    0.5, color=C_SQR_4)
    ax.bar(x[2], sqr_4x8_alpha, 0.5, bottom=sqr_4x8_pe, color=C_ALPHA_4)
    ax.bar(x[3], sqr_8x8_pe,    0.5, color=C_SQR_4)
    ax.bar(x[3], sqr_8x8_alpha, 0.5, bottom=sqr_8x8_pe, color=C_ALPHA_4)

    ref = bas_4x8
    def pct(v): return (v - ref) / ref * 100.0
    ax.text(x[0], bas_4x8 * 1.01, baseline_fmt.format(v=bas_4x8),    ha="center", va="bottom", fontsize=9)
    ax.text(x[1], bas_8x8 * 1.01, f"{pct(bas_8x8):+.1f}%", ha="center", va="bottom", fontsize=9)
    ax.text(x[2], sqr_4x8 * 1.01, f"{pct(sqr_4x8):+.1f}%", ha="center", va="bottom", fontsize=9)
    ax.text(x[3], sqr_8x8 * 1.01, f"{pct(sqr_8x8):+.1f}%", ha="center", va="bottom", fontsize=9)

    ax.set_ylim(0, max(bas_4x8, bas_8x8, sqr_4x8, sqr_8x8) * 1.40)
    ax.set_xticks(x)
    ax.set_xticklabels(["Baseline 4x8", "Baseline 8x8", "Square 4x8", "Square 8x8"])
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.legend(handles=[
        Patch(color=C_BAS_4,   label=f"PE Baseline 4x8 ×{n*n}"),
        Patch(color=C_SQR_4,   label=f"PE ×{n*n}"),
        Patch(color=C_ALPHA_4, label=f"Alpha ×{n}"),
    ])
    plt.tight_layout()
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"  wrote {out}")


def plot_ai_core_freq(d: dict, out: Path) -> None:
    bas_4x8 = d["Baseline 4x8"]["freq_ghz"]
    bas_8x8 = d["Baseline 8x8"]["freq_ghz"]
    sqr_4x8 = min(d["Square 4x8 SC"]["freq_ghz"], d["Square 4x8 Alpha Squared"]["freq_ghz"])
    sqr_8x8 = min(d["Square 8x8"]["freq_ghz"],    d["Square 8x8 Alpha Squared"]["freq_ghz"])

    x = np.arange(4)
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(x[0], bas_4x8, 0.5, color=C_BAS_4)
    ax.bar(x[1], bas_8x8, 0.5, color=C_SQR_4)
    ax.bar(x[2], sqr_4x8, 0.5, color=C_SQR_4)
    ax.bar(x[3], sqr_8x8, 0.5, color=C_SQR_4)

    ref = bas_4x8
    def pct(v): return (v - ref) / ref * 100.0
    ax.text(x[0], bas_4x8 * 1.01, f"{bas_4x8:.3f} GHz",    ha="center", va="bottom", fontsize=9)
    ax.text(x[1], bas_8x8 * 1.01, f"{pct(bas_8x8):+.1f}%", ha="center", va="bottom", fontsize=9)
    ax.text(x[2], sqr_4x8 * 1.01, f"{pct(sqr_4x8):+.1f}%", ha="center", va="bottom", fontsize=9)
    ax.text(x[3], sqr_8x8 * 1.01, f"{pct(sqr_8x8):+.1f}%", ha="center", va="bottom", fontsize=9)

    ax.set_ylim(0, max(bas_4x8, bas_8x8, sqr_4x8, sqr_8x8) * 1.40)
    ax.set_xticks(x)
    ax.set_xticklabels(["Baseline 4x8", "Baseline 8x8", "Square 4x8", "Square 8x8"])
    ax.set_ylabel("f_max (GHz)")
    ax.set_title("Frequency Analysis: AI-Core Level")
    ax.legend(handles=[
        Patch(color=C_BAS_4, label="PE Baseline 4x8"),
        Patch(color=C_SQR_4, label="PE (f_max = min over components)"),
    ])
    plt.tight_layout()
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"  wrote {out}")


def main() -> int:
    root = repo_root()
    xlsx = root / "doc" / "data" / "regres" / "results.xlsx"
    charts = root / "doc" / "charts" / "regres"
    charts.mkdir(parents=True, exist_ok=True)

    d = load_xlsx(xlsx)

    plot_pe_level(d, "area_um2", "Area Analysis: PE Level",      "Area (µm²)",      "{v:.1f} µm²",  charts / "area_pe_level.png")
    plot_pe_level(d, "freq_ghz", "Frequency Analysis: PE Level", "Freq Max (GHz)",  "{v:.3f} GHz",  charts / "freq_pe_level.png")
    plot_pe_level(d, "power_mw", "Power Analysis: PE Level",     "Power (mW)",      "{v:.2f} mW",   charts / "power_pe_level.png")

    for n in (8, 16):
        plot_ai_core_stacked(d, "area_um2", n,
            f"Area Analysis: AI-Core Level ({n}x{n})", "Area (µm²)",
            "{v:.0f} µm²", include_alpha_sum=False,
            out=charts / f"area_ai_core_level_{n}.png")
        plot_ai_core_stacked(d, "power_mw", n,
            f"Power Analysis: AI-Core Level ({n}x{n})", "Power (mW)",
            "{v:.1f} mW", include_alpha_sum=True,
            out=charts / f"power_ai_core_level_{n}.png")

    plot_ai_core_freq(d, charts / "freq_ai_core_level.png")
    return 0


if __name__ == "__main__":
    sys.exit(main())
