#!/usr/bin/env python3

# -----------------------------------------------------------------------------
# Author: Simone Machetti
#
# Collate the 32 DPA reports from scripts/flow/a_sweep/run.py and write
# doc/data/a_sweep/results.xlsx (8 sheets — see scripts/flow/a_sweep/README.md).
# -----------------------------------------------------------------------------

from __future__ import annotations

import os
import sys
from pathlib import Path

from openpyxl import Workbook

A_VALUES = (0, 1, 2, 3, 4, 5, 6, 7)
B_FIXED = 127
SUFFIX = "asw"


def repo_root() -> Path:
    code_home = os.environ.get("CODE_HOME")
    if not code_home:
        sys.exit("CODE_HOME is not set; run `source sourceme.sh` first.")
    return Path(code_home) / "ai-core" / "projects" / "ai-core"


def read_power_w(rpt: Path) -> float:
    for line in rpt.read_text().splitlines():
        if line.startswith("Total"):
            return float(line.split()[4])
    raise RuntimeError(f"No 'Total' row in {rpt}")


def to_uw(w: float) -> float:
    return w * 1e6


def load_b_arch(imp: Path, slug: str) -> dict[int, float]:
    out = {}
    for a in A_VALUES:
        rpt = imp / f"{slug}_{SUFFIX}_a{a}_b{B_FIXED}_dpa" / "report" / "power_summary.rpt"
        out[a] = round(to_uw(read_power_w(rpt)), 1)
    return out


def load_alpha(imp: Path, slug: str) -> dict[int, float]:
    out = {}
    for a in A_VALUES:
        rpt = imp / f"{slug}_{SUFFIX}_a{a}_dpa" / "report" / "power_summary.rpt"
        out[a] = round(to_uw(read_power_w(rpt)), 1)
    return out


def write_power_sheet(wb: Workbook, name: str, data: dict[int, float]) -> None:
    ws = wb.create_sheet(name)
    ws.append(["A", "power_uw"])
    for a in A_VALUES:
        ws.append([a, data[a]])


def write_config(wb: Workbook, name: str,
                 bas: dict, sqr_sc: dict, a_sqr: dict, a_sum: dict,
                 bas_mul: int, sqr_mul: int,
                 alpha_sqr_mul: int, alpha_sum_mul: int) -> None:
    ws = wb.create_sheet(name)
    ws.append(["A", "baseline_mw", "square_mw", "delta_mw", "improvement_pct"])
    for a in A_VALUES:
        base_total_uw = bas[a] * bas_mul
        sqr_total_uw = (sqr_sc[a] * sqr_mul
                        + a_sqr[a] * alpha_sqr_mul
                        + a_sum[a] * alpha_sum_mul)
        delta_uw = base_total_uw - sqr_total_uw
        imp_pct = 100.0 * delta_uw / base_total_uw if base_total_uw > 0 else 0.0
        ws.append([
            a,
            round(base_total_uw / 1000, 2),
            round(sqr_total_uw / 1000, 2),
            round(delta_uw / 1000, 2),
            round(imp_pct, 2),
        ])


def main() -> int:
    root = repo_root()
    imp = root / "imp"
    out = root / "doc" / "data" / "a_sweep" / "results.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    bas    = load_b_arch(imp, "bas_4x8")
    sqr_sc = load_b_arch(imp, "sqr_4x8_sc")
    a_sum  = load_alpha(imp,  "sqr_4x8_alpha_sum")
    a_sqr  = load_alpha(imp,  "sqr_4x8_alpha_sqr")

    wb = Workbook()
    wb.remove(wb.active)
    write_power_sheet(wb, "power_bas_4x8",    bas)
    write_power_sheet(wb, "power_sqr_4x8_sc", sqr_sc)
    write_power_sheet(wb, "power_alpha_sum",  a_sum)
    write_power_sheet(wb, "power_alpha_sqr",  a_sqr)
    write_config(wb, "config_1", bas, sqr_sc, a_sqr, a_sum,
                 bas_mul=256, sqr_mul=256, alpha_sqr_mul=64, alpha_sum_mul=0)
    write_config(wb, "config_2", bas, sqr_sc, a_sqr, a_sum,
                 bas_mul=256, sqr_mul=256, alpha_sqr_mul=64, alpha_sum_mul=48)
    write_config(wb, "config_3", bas, sqr_sc, a_sqr, a_sum,
                 bas_mul=64,  sqr_mul=64,  alpha_sqr_mul=32, alpha_sum_mul=24)
    write_config(wb, "config_4", bas, sqr_sc, a_sqr, a_sum,
                 bas_mul=128, sqr_mul=128, alpha_sqr_mul=48, alpha_sum_mul=40)

    wb.save(out)
    print(f"  wrote {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
