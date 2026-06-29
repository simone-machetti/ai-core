#!/usr/bin/env python3

# -----------------------------------------------------------------------------
# Author: Simone Machetti
#
# Read doc/data/dyn_range/results.xlsx and write 6 charts under
# doc/charts/dyn_range/.
# See scripts/flow/dyn_range/README.md for chart definitions.
# -----------------------------------------------------------------------------

from __future__ import annotations

import os
import sys
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401  (registers 3d projection)
from openpyxl import load_workbook

A_VALUES = (1, 3, 5, 7)
B_VALUES = (1, 19, 37, 55, 73, 91, 109, 127)


def repo_root() -> Path:
    code_home = os.environ.get("CODE_HOME")
    if not code_home:
        sys.exit("CODE_HOME is not set; run `source sourceme.sh` first.")
    return Path(code_home) / "ai-core" / "projects" / "ai-core"


def load_2d(wb, sheet: str) -> np.ndarray:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # row 0 is header ("A \ B", B values...); subsequent rows are (A, b1, b2, ...)
    data = np.zeros((len(A_VALUES), len(B_VALUES)))
    for r in rows[1:]:
        a = int(r[0])
        data[A_VALUES.index(a)] = [float(v) for v in r[1:]]
    return data


def load_1d(wb, sheet: str) -> np.ndarray:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    data = np.zeros(len(A_VALUES))
    for r in rows[1:]:
        a = int(r[0])
        data[A_VALUES.index(a)] = float(r[1])
    return data


def heatmap(data: np.ndarray, title: str, cbar_label: str,
            fmt: str, out: Path) -> None:
    vmax = float(np.max(np.abs(data)))
    fig, ax = plt.subplots(figsize=(9, 4.5))
    im = ax.imshow(data, cmap="RdYlGn", vmin=-vmax, vmax=vmax, aspect="auto")
    ax.set_xticks(np.arange(len(B_VALUES)))
    ax.set_yticks(np.arange(len(A_VALUES)))
    ax.set_xticklabels(B_VALUES)
    ax.set_yticklabels(A_VALUES)
    ax.set_xlabel("MAX_VAL_B")
    ax.set_ylabel("MAX_VAL_A")
    ax.set_title(title)
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            v = data[i, j]
            color = "white" if abs(v) > vmax * 0.55 else "black"
            ax.text(j, i, fmt.format(v=v), ha="center", va="center",
                    fontsize=9, color=color)
    cbar = fig.colorbar(im, ax=ax)
    cbar.set_label(cbar_label)
    plt.tight_layout()
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"  wrote {out}")


def compute_abs_mw(bas: np.ndarray, sqr_sc: np.ndarray,
                   a_sum: np.ndarray, a_sqr: np.ndarray) -> np.ndarray:
    base_total = bas * 256
    sqr_total  = sqr_sc * 256 + a_sqr[:, None] * 64 + a_sum[:, None] * 48
    return (base_total - sqr_total) / 1000.0  # µW → mW


def plot_abs_lines(abs_mw: np.ndarray, out: Path) -> None:
    colors  = ["#d73027", "#006837", "#66bd63", "#a6d96a"]
    markers = ["o", "s", "^", "D"]
    fig, ax = plt.subplots(figsize=(9, 5))
    for i, a in enumerate(A_VALUES):
        ax.plot(B_VALUES, abs_mw[i], color=colors[i], marker=markers[i],
                linewidth=2, markersize=7, label=f"MAX_VAL_A = {a}")
    avg = abs_mw.mean(axis=0)
    ax.plot(B_VALUES, avg, color="black", linestyle="--", linewidth=2.2,
            marker="x", markersize=8, label="Average across A")
    ax.axhline(0, color="black", linewidth=0.8, linestyle=":", alpha=0.5)
    ax.set_xticks(B_VALUES)
    ax.set_xlabel("MAX_VAL_B")
    ax.set_ylabel("Power saved (mW)  —  positive = Square wins")
    ax.set_title("Configuration 2 — Baseline − Square (mW) with both alpha arms")
    ax.grid(True, alpha=0.3)
    ax.legend(loc="best")
    plt.tight_layout()
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"  wrote {out}")


def plot_3d(abs_mw: np.ndarray, out: Path) -> None:
    fig = plt.figure(figsize=(16, 10))
    ax = fig.add_subplot(111, projection="3d")
    na, nb = len(A_VALUES), len(B_VALUES)
    xi = np.arange(na); yi = np.arange(nb)
    xx, yy = np.meshgrid(xi, yi, indexing="ij")
    dx, dy = 0.45, 0.45
    x = xx.ravel() - dx / 2
    y = yy.ravel() - dy / 2
    h = abs_mw.ravel()
    z0 = np.where(h >= 0, 0.0, h)
    dz = np.abs(h)
    colors = np.where(h >= 0, "#2ca02c", "#d62728")
    ax.bar3d(x, y, z0, dx, dy, dz,
             color=colors, edgecolor="black", linewidth=0.3, shade=True)
    ax.set_xticks(xi); ax.set_yticks(yi)
    ax.set_xticklabels(A_VALUES); ax.set_yticklabels(B_VALUES)
    ax.invert_xaxis()
    ax.set_xlabel("MAX_VAL_A"); ax.set_ylabel("MAX_VAL_B")
    ax.set_zlabel("Power win, Baseline - Square (mW)")
    ax.set_title("Absolute power win of Square 4x8 vs Baseline 4x8 - 16x16 PE Array")
    ax.legend(handles=[
        Patch(facecolor="#2ca02c", edgecolor="black", label="Square wins (positive mW saved)"),
        Patch(facecolor="#d62728", edgecolor="black", label="Baseline wins (negative mW)"),
    ], loc="upper left", fontsize=11)
    ax.view_init(elev=28, azim=-55)
    plt.tight_layout()
    plt.savefig(out, dpi=200)
    plt.close(fig)
    print(f"  wrote {out}")


def main() -> int:
    root = repo_root()
    xlsx = root / "doc" / "data" / "dyn_range" / "results.xlsx"
    charts = root / "doc" / "charts" / "dyn_range"
    charts.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx, data_only=True)
    bas    = load_2d(wb, "power_bas_4x8")
    sqr_sc = load_2d(wb, "power_sqr_4x8_sc")
    a_sum  = load_1d(wb, "power_alpha_sum")
    a_sqr  = load_1d(wb, "power_alpha_sqr")
    cfg_2 = load_2d(wb, "config_2")
    cfg_3 = load_2d(wb, "config_3")
    cfg_4 = load_2d(wb, "config_4")

    heatmap(cfg_2,
            "Configuration 2 — Square vs Baseline (%) with both alpha arms",
            "Improvement (%)  —  positive = Square wins",
            "{v:+.2f}",
            charts / "config_2.png")
    heatmap(cfg_3,
            "Configuration 3 — Square vs Baseline (%) — 8x8 PE array",
            "Improvement (%)  —  positive = Square wins",
            "{v:+.2f}",
            charts / "config_3.png")
    heatmap(cfg_4,
            "Configuration 4 — Square vs Baseline (%) — 8x16 PE array",
            "Improvement (%)  —  positive = Square wins",
            "{v:+.2f}",
            charts / "config_4.png")

    abs_mw = compute_abs_mw(bas, sqr_sc, a_sum, a_sqr)
    heatmap(abs_mw,
            "Configuration 2 — Baseline − Square (mW) with both alpha arms",
            "Power saved (mW)  —  positive = Square wins",
            "{v:+.1f}",
            charts / "config_2_abs.png")
    plot_abs_lines(abs_mw, charts / "config_2_abs_lines.png")
    plot_3d(abs_mw, charts / "config_2_3d.png")
    return 0


if __name__ == "__main__":
    sys.exit(main())
